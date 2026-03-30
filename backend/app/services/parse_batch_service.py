"""Background batch parsing service for large-scale TTP jobs."""
from __future__ import annotations

import json
import os
import shutil
import threading
import time
import tempfile
import traceback
import uuid
import zipfile
from concurrent.futures import FIRST_COMPLETED, Future, ProcessPoolExecutor, ThreadPoolExecutor, wait
from multiprocessing import get_context
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook

from .ttp_service import TTPService


def _env_int(name: str, default: int, minimum: int = 1) -> int:
    """Read a bounded integer from the environment."""
    raw_value = os.getenv(name)
    if raw_value is None:
        return default
    try:
        return max(minimum, int(raw_value))
    except (TypeError, ValueError):
        return default


DEFAULT_BATCH_ROOT = Path(tempfile.gettempdir()) / "ttp_batch_jobs"
SUPPORTED_EXTENSIONS = {".txt", ".log", ".cfg", ".conf", ".json", ".xml", ".yaml", ".yml"}
RESULTS_PAGE_DEFAULT = 50
RESULTS_PAGE_MAX = 200
RESULTS_PREVIEW_LIMIT = 100
EXECUTOR_KIND = (os.getenv("TTP_BATCH_EXECUTOR") or "process").strip().lower()
MAX_PARSE_WORKERS = _env_int(
    "TTP_BATCH_MAX_WORKERS",
    min(32, max(4, os.cpu_count() or 4)),
)
MAX_PENDING_FUTURES = _env_int(
    "TTP_BATCH_MAX_PENDING_FUTURES",
    MAX_PARSE_WORKERS * 4,
    minimum=MAX_PARSE_WORKERS,
)
SCAN_PROGRESS_FLUSH_INTERVAL_MS = 250
PARSE_PROGRESS_FLUSH_INTERVAL_MS = _env_int(
    "TTP_BATCH_PROGRESS_FLUSH_INTERVAL_MS",
    500,
    minimum=100,
)
RESULTS_FLUSH_EVERY = _env_int("TTP_BATCH_RESULTS_FLUSH_EVERY", 100, minimum=1)
STATE_FILE_RETRY_COUNT = _env_int("TTP_BATCH_STATE_RETRY_COUNT", 8, minimum=1)
STATE_FILE_RETRY_DELAY_MS = _env_int("TTP_BATCH_STATE_RETRY_DELAY_MS", 25, minimum=1)


def _current_timestamp() -> int:
    """Return the current timestamp in milliseconds."""
    return time.time_ns() // 1_000_000


def _safe_name(name: str, fallback: str) -> str:
    """Return a filesystem-safe file name."""
    cleaned = "".join(char if char.isalnum() or char in {"-", "_", "."} else "_" for char in name)
    cleaned = cleaned.strip("._")
    return cleaned or fallback


def _decode_text_bytes(content: bytes) -> str:
    """Decode bytes using utf-8 with latin-1 fallback."""
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def _extract_variable_names_from_template(template_text: str) -> list[str]:
    """Extract variable names in definition order from TTP template text."""
    seen: set[str] = set()
    names: list[str] = []
    for match in re.finditer(r"\{\{\s*(\w+)\s*(?:\|[^}]*)?\}\}", template_text):
        name = match.group(1)
        if name not in seen:
            seen.add(name)
            names.append(name)
    return names


def _stringify_excel_value(value: Any) -> str:
    """Convert non-scalar values into a stable Excel-safe string."""
    if isinstance(value, str):
        return value

    json_value = json.dumps(value, ensure_ascii=False)
    return json_value if json_value is not None else str(value)


def _to_excel_cell_value(value: Any) -> str | int | float | bool | None:
    """Normalize a Python value into a scalar Excel cell value."""
    if value is None:
        return None

    if isinstance(value, (str, int, float, bool)):
        return value

    return _stringify_excel_value(value)


def _combine_excel_rows(
    current_rows: list[dict[str, str | int | float | bool | None]],
    next_rows: list[dict[str, str | int | float | bool | None]],
) -> list[dict[str, str | int | float | bool | None]]:
    """Produce the cartesian merge of sibling row sets."""
    combined: list[dict[str, str | int | float | bool | None]] = []
    for current_row in current_rows:
        for next_row in next_rows:
            combined.append({**current_row, **next_row})
    return combined


def _flatten_result_to_rows(
    value: Any,
    path: tuple[str, ...] = (),
) -> list[dict[str, str | int | float | bool | None]]:
    """Flatten a nested result into tabular rows."""
    if isinstance(value, list):
        if not value:
            return [{}]
        rows: list[dict[str, str | int | float | bool | None]] = []
        for item in value:
            rows.extend(_flatten_result_to_rows(item, path))
        return rows

    if isinstance(value, dict):
        if not value:
            return [{}]

        rows: list[dict[str, str | int | float | bool | None]] = [{}]
        for key, child_value in value.items():
            rows = _combine_excel_rows(rows, _flatten_result_to_rows(child_value, (*path, str(key))))
        return rows

    if not path:
        return [{}]

    return [{".".join(path): _to_excel_cell_value(value)}]


def _build_result_excel_rows(result: Any) -> list[dict[str, str | int | float | bool | None]]:
    """Build worksheet rows for one parsed result payload."""
    if result is None:
        return []

    rows = _flatten_result_to_rows(result)
    has_structured_columns = any(bool(row) for row in rows)
    if has_structured_columns:
        return rows

    return []


def _sanitize_excel_sheet_name(name: str) -> str:
    """Return an Excel-compatible worksheet name."""
    trimmed = (name or "").strip() or "Template"
    sanitized = "".join("_" if char in {":", "\\", "/", "?", "*", "[", "]"} else char for char in trimmed)
    sanitized = sanitized.replace("'", "").strip()
    return (sanitized or "Template")[:31]


def _build_unique_excel_sheet_name(template_name: str, used_names: set[str]) -> str:
    """Ensure worksheet names stay unique within a workbook."""
    base_name = _sanitize_excel_sheet_name(template_name)
    if base_name not in used_names:
        used_names.add(base_name)
        return base_name

    suffix = 2
    while True:
        label = f" ({suffix})"
        candidate = f"{base_name[: 31 - len(label)]}{label}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        suffix += 1


def _is_supported_path(path: str) -> bool:
    """Return True when the path extension is parseable text."""
    suffix = Path(path).suffix.lower()
    return suffix in SUPPORTED_EXTENSIONS


def _safe_relative_path(path: str, fallback_prefix: str) -> Path:
    """Normalize an archive entry path into a safe relative path."""
    candidate = Path(path.replace("\\", "/"))
    safe_parts = [
        _safe_name(part, fallback_prefix)
        for part in candidate.parts
        if part not in {"", ".", ".."}
    ]
    if not safe_parts:
        safe_parts = [fallback_prefix]
    return Path(*safe_parts)


def _work_item_display_name(work_item: dict[str, Any]) -> str:
    """Return a stable display name for a plain file or archive entry."""
    if work_item["kind"] == "plain_file":
        return work_item["file_name"]
    return f"{work_item['archive_name']}::{work_item['entry_name']}"


def _load_work_item_text_payload(work_item: dict[str, Any]) -> tuple[str, str]:
    """Load text content and display name for a serializable work item."""
    if work_item["kind"] == "plain_file":
        content = Path(work_item["path"]).read_bytes()
        return work_item["file_name"], _decode_text_bytes(content)

    archive_path = Path(work_item["archive_path"])
    with zipfile.ZipFile(archive_path) as archive:
        content = archive.read(work_item["entry_name"])
    return _work_item_display_name(work_item), _decode_text_bytes(content)


def _parse_task(
    file_name: str,
    template_id: str,
    template_name: str,
    template_text: str,
    data: str,
    variable_names: list[str] | None = None,
) -> dict[str, Any]:
    """Run one parse task and return a serializable result row."""
    started_at = _current_timestamp()
    parsed = TTPService.parse(
        data=data,
        template=template_text,
        include_csv=False,
        include_checkup=False,
        variable_names=variable_names,
    )
    finished_at = _current_timestamp()
    return {
        "file_name": file_name,
        "template_id": template_id,
        "template_name": template_name,
        "success": parsed["success"],
        "result": parsed.get("result"),
        "csv_result": parsed.get("csv_result"),
        "checkup_csv_result": parsed.get("checkup_csv_result"),
        "error": parsed.get("error"),
        "error_type": parsed.get("error_type"),
        "started_at": started_at,
        "finished_at": finished_at,
    }


def _parse_work_item(
    work_item: dict[str, Any],
    templates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Load one work item and parse it against all selected templates."""
    file_name, data = _load_work_item_text_payload(work_item)
    return [
        _parse_task(
            file_name=file_name,
            template_id=template["id"],
            template_name=template["name"],
            template_text=template["template"],
            data=data,
            variable_names=template.get("variable_names"),
        )
        for template in templates
    ]


class ParseBatchService:
    """Manage background batch parsing jobs and artifacts."""

    _lock = threading.Lock()
    _job_threads: dict[str, threading.Thread] = {}
    _state_locks: dict[str, threading.Lock] = {}

    @classmethod
    def get_root_path(cls, root_path: str | Path | None = None) -> Path:
        """Resolve the batch job root directory."""
        configured_path = root_path or os.getenv("TTP_BATCH_JOBS_PATH")
        path = Path(configured_path) if configured_path else DEFAULT_BATCH_ROOT
        if not path.is_absolute():
            path = Path.cwd() / path
        return path

    @classmethod
    def initialize(cls, root_path: str | Path | None = None) -> Path:
        """Ensure the batch job root directory exists."""
        path = cls.get_root_path(root_path)
        path.mkdir(parents=True, exist_ok=True)
        return path

    @classmethod
    def _job_dir(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls.initialize(root_path) / job_id

    @classmethod
    def _state_path(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls._job_dir(job_id, root_path) / "status.json"

    @classmethod
    def _results_path(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls._job_dir(job_id, root_path) / "results.jsonl"

    @classmethod
    def _errors_path(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls._job_dir(job_id, root_path) / "errors.jsonl"

    @classmethod
    def _excel_path(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls._job_dir(job_id, root_path) / "results.xlsx"

    @classmethod
    def _summary_path(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls._job_dir(job_id, root_path) / "summary.json"

    @classmethod
    def _extracted_dir(cls, job_id: str, root_path: str | Path | None = None) -> Path:
        return cls._job_dir(job_id, root_path) / "extracted"

    @classmethod
    def _get_state_lock(cls, job_id: str) -> threading.Lock:
        with cls._lock:
            lock = cls._state_locks.get(job_id)
            if lock is None:
                lock = threading.Lock()
                cls._state_locks[job_id] = lock
            return lock

    @classmethod
    def _load_state(cls, job_id: str, root_path: str | Path | None = None) -> dict[str, Any]:
        state_path = cls._state_path(job_id, root_path)
        if not state_path.exists():
            raise FileNotFoundError(f"Batch parse job not found: {job_id}")
        state_lock = cls._get_state_lock(job_id)
        last_error: Exception | None = None
        for attempt in range(STATE_FILE_RETRY_COUNT):
            try:
                with state_lock:
                    return json.loads(state_path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError) as exc:
                last_error = exc
                if attempt == STATE_FILE_RETRY_COUNT - 1:
                    break
                time.sleep(STATE_FILE_RETRY_DELAY_MS / 1000)
        if last_error is not None:
            raise last_error
        raise FileNotFoundError(f"Batch parse job not found: {job_id}")

    @classmethod
    def _write_state(
        cls,
        job_id: str,
        state: dict[str, Any],
        root_path: str | Path | None = None,
    ) -> None:
        state["updated_at"] = _current_timestamp()
        state_path = cls._state_path(job_id, root_path)
        state_lock = cls._get_state_lock(job_id)
        payload = json.dumps(state, ensure_ascii=False, indent=2)
        last_error: Exception | None = None

        for attempt in range(STATE_FILE_RETRY_COUNT):
            temp_state_path = state_path.with_name(f"{state_path.stem}.{threading.get_ident()}.{attempt}.tmp")
            try:
                with state_lock:
                    temp_state_path.write_text(payload, encoding="utf-8")
                    os.replace(temp_state_path, state_path)
                return
            except OSError as exc:
                last_error = exc
                try:
                    if temp_state_path.exists():
                        temp_state_path.unlink()
                except OSError:
                    pass
                if attempt == STATE_FILE_RETRY_COUNT - 1:
                    break
                time.sleep(STATE_FILE_RETRY_DELAY_MS / 1000)

        if last_error is not None:
            raise last_error

    @classmethod
    async def create_job(
        cls,
        templates: list[dict[str, str]],
        uploads: list[UploadFile],
        root_path: str | Path | None = None,
    ) -> dict[str, Any]:
        """Create a job, persist uploads, and start background processing."""
        if not templates:
            raise ValueError("At least one template is required")
        if not uploads:
            raise ValueError("At least one file upload is required")

        normalized_templates: list[dict[str, Any]] = []
        for template in templates:
            template_text = (template.get("template") or "").strip()
            if not template_text:
                continue
            normalized_templates.append(
                {
                    "id": template.get("id") or str(uuid.uuid4()),
                    "name": template.get("name") or "Unnamed Template",
                    "template": template_text,
                    "variable_names": template.get("variable_names"),
                }
            )

        if not normalized_templates:
            raise ValueError("No parseable templates were provided")

        job_id = str(uuid.uuid4())
        job_dir = cls._job_dir(job_id, root_path)
        uploads_dir = job_dir / "uploads"
        uploads_dir.mkdir(parents=True, exist_ok=True)

        persisted_uploads: list[dict[str, Any]] = []
        for index, upload in enumerate(uploads):
            upload_name = upload.filename or f"upload-{index + 1}"
            safe_name = _safe_name(upload_name, f"upload-{index + 1}")
            target_path = uploads_dir / f"{index:05d}-{safe_name}"
            with target_path.open("wb") as target_file:
                shutil.copyfileobj(upload.file, target_file)
            await upload.close()
            persisted_uploads.append(
                {
                    "original_name": upload_name,
                    "stored_name": target_path.name,
                    "path": str(target_path),
                    "is_archive": target_path.suffix.lower() == ".zip",
                    "size": target_path.stat().st_size,
                }
            )

        state = {
            "id": job_id,
            "status": "queued",
            "cancel_requested": False,
            "phase_message": "Waiting for background worker",
            "created_at": _current_timestamp(),
            "updated_at": _current_timestamp(),
            "started_at": None,
            "completed_at": None,
            "template_count": len(normalized_templates),
            "upload_count": len(persisted_uploads),
            "scanned_uploads": 0,
            "total_uploads": len(persisted_uploads),
            "processed_archive_entries": 0,
            "total_archive_entries": 0,
            "uploads": [
                {
                    "name": item["original_name"],
                    "size": item["size"],
                    "is_archive": item["is_archive"],
                }
                for item in persisted_uploads
            ],
            "discovered_file_count": 0,
            "skipped_file_count": 0,
            "total_tasks": 0,
            "completed_tasks": 0,
            "success_count": 0,
            "failure_count": 0,
            "preview_results": [],
            "recent_error": None,
            "artifacts": {
                "summary": None,
                "results": None,
                "errors": None,
                "excel": None,
            },
        }
        cls._write_state(job_id, state, root_path)

        worker = threading.Thread(
            target=cls._run_job,
            args=(job_id, normalized_templates, persisted_uploads, root_path),
            daemon=True,
            name=f"parse-batch-{job_id}",
        )
        with cls._lock:
            cls._job_threads[job_id] = worker
        worker.start()
        return cls.get_job(job_id, root_path)

    @classmethod
    def get_job(cls, job_id: str, root_path: str | Path | None = None) -> dict[str, Any]:
        """Return a job status payload with artifact URLs."""
        state = cls._load_state(job_id, root_path)
        if state.get("status") == "running":
            state["status"] = "parsing"
        state.setdefault("cancel_requested", False)
        state.setdefault("phase_message", "Batch job state loaded")
        state.setdefault("scanned_uploads", 0)
        state.setdefault("total_uploads", state.get("upload_count", 0))
        state.setdefault("processed_archive_entries", 0)
        state.setdefault("total_archive_entries", 0)
        state["artifact_urls"] = {
            key: (
                f"/api/parse/batch/jobs/{job_id}/artifacts/{key}"
                if state["artifacts"].get(key)
                else None
            )
            for key in state["artifacts"]
        }
        return state

    @classmethod
    def cancel_job(cls, job_id: str, root_path: str | Path | None = None) -> dict[str, Any]:
        """Request cancellation for a running batch parse job."""
        state = cls._load_state(job_id, root_path)
        if cls._is_terminal_status(state.get("status")):
            return cls.get_job(job_id, root_path)

        state["cancel_requested"] = True
        state["status"] = "cancel_requested"
        state["phase_message"] = "Stopping batch job"
        cls._write_state(job_id, state, root_path)
        return cls.get_job(job_id, root_path)

    @classmethod
    def get_results_page(
        cls,
        job_id: str,
        offset: int = 0,
        limit: int = RESULTS_PAGE_DEFAULT,
        root_path: str | Path | None = None,
    ) -> dict[str, Any]:
        """Return a paginated slice of batch parse results."""
        state = cls.get_job(job_id, root_path)
        normalized_offset = max(offset, 0)
        normalized_limit = min(max(limit, 1), RESULTS_PAGE_MAX)
        results_path = cls._results_path(job_id, root_path)

        items: list[dict[str, Any]] = []
        if results_path.exists():
            with results_path.open("r", encoding="utf-8") as handle:
                for line_index, line in enumerate(handle):
                    if line_index < normalized_offset:
                        continue
                    if len(items) >= normalized_limit:
                        break
                    items.append(json.loads(line))

        return {
            "job_id": job_id,
            "offset": normalized_offset,
            "limit": normalized_limit,
            "total": state["completed_tasks"],
            "items": items,
        }

    @classmethod
    def get_artifact_path(
        cls,
        job_id: str,
        artifact_name: str,
        root_path: str | Path | None = None,
    ) -> Path:
        """Return the on-disk artifact path for a known job artifact."""
        artifact_map = {
            "summary": cls._summary_path(job_id, root_path),
            "results": cls._results_path(job_id, root_path),
            "errors": cls._errors_path(job_id, root_path),
            "excel": cls._excel_path(job_id, root_path),
        }
        artifact_path = artifact_map.get(artifact_name)
        if artifact_path is None or not artifact_path.exists():
            raise FileNotFoundError(f"Artifact not found: {artifact_name}")
        return artifact_path

    @classmethod
    def _write_excel_artifact(
        cls,
        job_id: str,
        root_path: str | Path | None = None,
        templates: list[dict[str, Any]] | None = None,
    ) -> Path | None:
        """Build a result-only Excel workbook for a completed batch job."""
        results_path = cls._results_path(job_id, root_path)
        if not results_path.exists():
            return None

        template_var_order: dict[str, list[str]] = {}
        for t in (templates or []):
            tid = t.get("id", "")
            if not tid:
                continue
            var_names = t.get("variable_names")
            if var_names:
                template_var_order[tid] = var_names
            elif t.get("template"):
                template_var_order[tid] = _extract_variable_names_from_template(t["template"])

        grouped_rows: dict[str, dict[str, Any]] = {}
        with results_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                item = json.loads(line)
                template_key = str(item.get("template_id") or item.get("template_name") or "template")
                template_name = str(item.get("template_name") or "Unnamed Template")
                excel_rows = _build_result_excel_rows(item.get("result"))
                if not excel_rows:
                    continue

                if template_key not in grouped_rows:
                    grouped_rows[template_key] = {
                        "template_id": template_key,
                        "template_name": template_name,
                        "rows": list(excel_rows),
                    }
                    continue

                grouped_rows[template_key]["rows"].extend(excel_rows)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        used_sheet_names: set[str] = set()
        if not grouped_rows:
            workbook.create_sheet(title="Results")
        else:
            for template_data in grouped_rows.values():
                sheet_name = _build_unique_excel_sheet_name(template_data["template_name"], used_sheet_names)
                worksheet = workbook.create_sheet(title=sheet_name)
                rows = template_data["rows"]
                headers: list[str] = []
                seen_headers: set[str] = set()
                for row in rows:
                    for key in row:
                        if key not in seen_headers:
                            seen_headers.add(key)
                            headers.append(key)

                if not headers:
                    headers = ["value"]

                var_order = template_var_order.get(template_data.get("template_id", ""), [])
                if var_order:
                    ordered = [v for v in var_order if v in seen_headers]
                    ordered += [h for h in headers if h not in set(var_order)]
                    headers = ordered

                worksheet.append(headers)
                for row in rows:
                    worksheet.append([row.get(header) for header in headers])

        excel_path = cls._excel_path(job_id, root_path)
        workbook.save(excel_path)
        return excel_path

    @classmethod
    def _append_preview_result(cls, state: dict[str, Any], item: dict[str, Any]) -> None:
        if len(state["preview_results"]) < RESULTS_PREVIEW_LIMIT:
            state["preview_results"].append(item)

    @classmethod
    def _is_terminal_status(cls, status: str | None) -> bool:
        """Return True when a job status is terminal."""
        return status in {"completed", "failed", "cancelled"}

    @classmethod
    def _update_state_fields(
        cls,
        job_id: str,
        state: dict[str, Any],
        root_path: str | Path | None = None,
        **fields: Any,
    ) -> None:
        """Merge updated fields into the state and persist them."""
        state.update(fields)
        cls._write_state(job_id, state, root_path)

    @classmethod
    def _scan_uploads(
        cls,
        job_id: str,
        state: dict[str, Any],
        uploads: list[dict[str, Any]],
        root_path: str | Path | None = None,
    ) -> tuple[list[dict[str, Any]], int]:
        """Discover parseable files from uploaded raw files and zip archives."""
        work_items: list[dict[str, Any]] = []
        skipped_count = 0
        total_archive_entries = 0
        processed_archive_entries = 0
        scanned_uploads = 0
        last_flush_at = 0
        extracted_dir = cls._extracted_dir(job_id, root_path)

        def flush_progress(force: bool = False) -> None:
            nonlocal last_flush_at
            now = _current_timestamp()
            if not force and now - last_flush_at < SCAN_PROGRESS_FLUSH_INTERVAL_MS:
                return

            cls._update_state_fields(
                job_id,
                state,
                root_path,
                scanned_uploads=scanned_uploads,
                total_uploads=len(uploads),
                processed_archive_entries=processed_archive_entries,
                total_archive_entries=total_archive_entries,
                discovered_file_count=len(work_items),
                skipped_file_count=skipped_count,
            )
            last_flush_at = now

        for upload in uploads:
            if cls._refresh_cancel_requested(job_id, state, root_path):
                break
            upload_path = Path(upload["path"])
            if upload["is_archive"]:
                with zipfile.ZipFile(upload_path) as archive:
                    archive_entries = [info for info in archive.infolist() if not info.is_dir()]
                    total_archive_entries += len(archive_entries)
                    flush_progress(force=True)

                    for info in archive_entries:
                        if cls._refresh_cancel_requested(job_id, state, root_path):
                            break
                        processed_archive_entries += 1
                        if not _is_supported_path(info.filename):
                            skipped_count += 1
                            flush_progress()
                            continue
                        relative_path = _safe_relative_path(
                            info.filename,
                            f"entry-{processed_archive_entries}",
                        )
                        extracted_path = extracted_dir / upload_path.stem / relative_path
                        extracted_path.parent.mkdir(parents=True, exist_ok=True)
                        with archive.open(info, "r") as source, extracted_path.open("wb") as target:
                            shutil.copyfileobj(source, target)
                        work_items.append(
                            {
                                "kind": "plain_file",
                                "path": str(extracted_path),
                                "file_name": f"{upload['original_name']}::{info.filename}",
                            }
                        )
                        flush_progress()
                scanned_uploads += 1
                flush_progress(force=True)
                continue

            if not _is_supported_path(upload["original_name"]):
                skipped_count += 1
            else:
                work_items.append(
                    {
                        "kind": "plain_file",
                        "path": str(upload_path),
                        "file_name": upload["original_name"],
                    }
                )

            scanned_uploads += 1
            flush_progress(force=True)

        flush_progress(force=True)
        return work_items, skipped_count

    @classmethod
    def _refresh_cancel_requested(
        cls,
        job_id: str,
        state: dict[str, Any],
        root_path: str | Path | None = None,
    ) -> bool:
        """Refresh the cancel flag from disk and return the latest state."""
        if state.get("cancel_requested"):
            return True

        latest_state = cls._load_state(job_id, root_path)
        if latest_state.get("cancel_requested"):
            state["cancel_requested"] = True
            state["status"] = latest_state.get("status", state.get("status"))
            state["phase_message"] = latest_state.get("phase_message", state.get("phase_message"))
            return True
        return False

    @classmethod
    def _load_work_item_text(cls, work_item: dict[str, Any]) -> tuple[str, str]:
        """Load text content and display name for a single work item."""
        return _load_work_item_text_payload(work_item)

    @classmethod
    def _consume_completed_futures(
        cls,
        job_id: str,
        state: dict[str, Any],
        pending: set[Future[list[dict[str, Any]]]],
        future_meta: dict[Future[list[dict[str, Any]]], dict[str, Any]],
        results_handle,
        errors_handle,
        wait_for_one: bool = False,
    ) -> int:
        """Drain completed futures, persist results, and update job state."""
        if not pending:
            return 0

        done, not_done = wait(
            pending,
            timeout=None if wait_for_one else 0,
            return_when=FIRST_COMPLETED if wait_for_one else FIRST_COMPLETED,
        )
        if not done:
            return 0

        pending.clear()
        pending.update(not_done)
        completed_count = 0

        for future in done:
            meta = future_meta.pop(future, {})
            try:
                items = future.result()
            except Exception as exc:  # pragma: no cover - defensive guard
                items = [
                    {
                        "file_name": meta.get("file_name", "unknown"),
                        "template_id": template.get("id", ""),
                        "template_name": template.get("name", "Unknown Template"),
                        "success": False,
                        "result": None,
                        "csv_result": None,
                        "checkup_csv_result": None,
                        "error": str(exc),
                        "error_type": type(exc).__name__,
                        "started_at": None,
                        "finished_at": _current_timestamp(),
                    }
                    for template in meta.get("templates", [])
                ] or [
                    {
                        "file_name": meta.get("file_name", "unknown"),
                        "template_id": "",
                        "template_name": "Unknown Template",
                        "success": False,
                        "result": None,
                        "csv_result": None,
                        "checkup_csv_result": None,
                        "error": str(exc),
                        "error_type": type(exc).__name__,
                        "started_at": None,
                        "finished_at": _current_timestamp(),
                    }
                ]

            for item in items:
                compact_result = {
                    "file_name": item["file_name"],
                    "template_id": item["template_id"],
                    "template_name": item["template_name"],
                    "success": item["success"],
                    "result": item.get("result"),
                    "error": item.get("error"),
                    "error_type": item.get("error_type"),
                }
                results_handle.write(json.dumps(compact_result, ensure_ascii=False) + "\n")
                if not item["success"]:
                    errors_handle.write(
                        json.dumps(
                            {
                                "file_name": item["file_name"],
                                "template_name": item["template_name"],
                                "error": item.get("error"),
                                "error_type": item.get("error_type"),
                            },
                            ensure_ascii=False,
                        )
                        + "\n"
                    )

                preview_item = {
                    "file_name": item["file_name"],
                    "template_name": item["template_name"],
                    "success": item["success"],
                    "error": item.get("error"),
                    "error_type": item.get("error_type"),
                }
                cls._append_preview_result(state, preview_item)

                state["completed_tasks"] += 1
                if item["success"]:
                    state["success_count"] += 1
                else:
                    state["failure_count"] += 1
                    state["recent_error"] = preview_item
                completed_count += 1

        return completed_count

    @classmethod
    def _finalize_summary(
        cls,
        job_id: str,
        state: dict[str, Any],
        templates: list[dict[str, str]],
        root_path: str | Path | None = None,
    ) -> None:
        """Write the final summary artifact for a job."""
        summary = {
            "job": {
                "id": job_id,
                "status": state["status"],
                "phase_message": state["phase_message"],
                "created_at": state["created_at"],
                "started_at": state["started_at"],
                "completed_at": state["completed_at"],
            },
            "templates": [
                {
                    "id": template["id"],
                    "name": template["name"],
                }
                for template in templates
            ],
            "counts": {
                "upload_count": state["upload_count"],
                "scanned_uploads": state["scanned_uploads"],
                "total_uploads": state["total_uploads"],
                "processed_archive_entries": state["processed_archive_entries"],
                "total_archive_entries": state["total_archive_entries"],
                "discovered_file_count": state["discovered_file_count"],
                "skipped_file_count": state["skipped_file_count"],
                "template_count": state["template_count"],
                "total_tasks": state["total_tasks"],
                "completed_tasks": state["completed_tasks"],
                "success_count": state["success_count"],
                "failure_count": state["failure_count"],
            },
            "preview_results": state["preview_results"],
            "recent_error": state["recent_error"],
        }
        cls._summary_path(job_id, root_path).write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @classmethod
    def _mark_cancelled(
        cls,
        job_id: str,
        state: dict[str, Any],
        templates: list[dict[str, str]],
        root_path: str | Path | None = None,
    ) -> None:
        """Persist a cancelled job state and summary."""
        state["cancel_requested"] = True
        state["status"] = "cancelled"
        state["phase_message"] = "Batch parse cancelled"
        state["completed_at"] = _current_timestamp()
        state["artifacts"] = {
            "summary": cls._summary_path(job_id, root_path).name,
            "results": cls._results_path(job_id, root_path).name if cls._results_path(job_id, root_path).exists() else None,
            "errors": cls._errors_path(job_id, root_path).name if cls._errors_path(job_id, root_path).exists() else None,
            "excel": cls._excel_path(job_id, root_path).name if cls._excel_path(job_id, root_path).exists() else None,
        }
        cls._finalize_summary(job_id, state, templates, root_path)
        cls._write_state(job_id, state, root_path)

    @classmethod
    def _create_executor(cls):
        """Create the configured executor for batch parse work."""
        if EXECUTOR_KIND == "thread":
            return ThreadPoolExecutor(max_workers=MAX_PARSE_WORKERS)
        try:
            return ProcessPoolExecutor(
                max_workers=MAX_PARSE_WORKERS,
                mp_context=get_context("spawn"),
            )
        except (OSError, PermissionError, ValueError):
            return ThreadPoolExecutor(max_workers=MAX_PARSE_WORKERS)

    @classmethod
    def _run_job(
        cls,
        job_id: str,
        templates: list[dict[str, str]],
        uploads: list[dict[str, Any]],
        root_path: str | Path | None = None,
    ) -> None:
        """Process a batch parse job in the background."""
        state = cls._load_state(job_id, root_path)
        if state.get("cancel_requested"):
            cls._mark_cancelled(job_id, state, templates, root_path)
            return
        cls._update_state_fields(
            job_id,
            state,
            root_path,
            status="scanning",
            phase_message="Scanning uploads",
            started_at=_current_timestamp(),
        )

        try:
            work_items, skipped_count = cls._scan_uploads(job_id, state, uploads, root_path)
            if state.get("cancel_requested"):
                cls._mark_cancelled(job_id, state, templates, root_path)
                return
            cls._update_state_fields(
                job_id,
                state,
                root_path,
                status="parsing",
                phase_message="Parsing files",
                discovered_file_count=len(work_items),
                skipped_file_count=skipped_count,
                total_tasks=len(work_items) * len(templates),
            )

            results_path = cls._results_path(job_id, root_path)
            errors_path = cls._errors_path(job_id, root_path)

            with results_path.open("w", encoding="utf-8") as results_handle, errors_path.open(
                "w", encoding="utf-8"
            ) as errors_handle, cls._create_executor() as executor:
                pending: set[Future[list[dict[str, Any]]]] = set()
                future_meta: dict[Future[list[dict[str, Any]]], dict[str, Any]] = {}
                completed_since_flush = 0
                last_progress_flush_at = _current_timestamp()

                def flush_parse_progress(force: bool = False) -> None:
                    nonlocal completed_since_flush, last_progress_flush_at
                    now = _current_timestamp()
                    if not force:
                        if completed_since_flush < RESULTS_FLUSH_EVERY:
                            if now - last_progress_flush_at < PARSE_PROGRESS_FLUSH_INTERVAL_MS:
                                return

                    results_handle.flush()
                    errors_handle.flush()
                    cls._write_state(job_id, state, root_path)
                    completed_since_flush = 0
                    last_progress_flush_at = now

                for work_item in work_items:
                    if cls._refresh_cancel_requested(job_id, state, root_path):
                        break
                    future = executor.submit(_parse_work_item, work_item, templates)
                    pending.add(future)
                    future_meta[future] = {
                        "file_name": _work_item_display_name(work_item),
                        "templates": [{"id": template["id"], "name": template["name"]} for template in templates],
                    }

                    if len(pending) >= MAX_PENDING_FUTURES:
                        completed_since_flush += cls._consume_completed_futures(
                            job_id,
                            state,
                            pending,
                            future_meta,
                            results_handle,
                            errors_handle,
                            wait_for_one=True,
                        )
                        flush_parse_progress()

                while pending and not state.get("cancel_requested"):
                    completed_since_flush += cls._consume_completed_futures(
                        job_id,
                        state,
                        pending,
                        future_meta,
                        results_handle,
                        errors_handle,
                        wait_for_one=True,
                    )
                    flush_parse_progress(force=not pending)
                    cls._refresh_cancel_requested(job_id, state, root_path)

                flush_parse_progress(force=True)

                if state.get("cancel_requested"):
                    for future in list(pending):
                        future.cancel()
                    flush_parse_progress(force=True)
                    cls._mark_cancelled(job_id, state, templates, root_path)
                    return

            excel_path = cls._write_excel_artifact(job_id, root_path, templates=templates)

            state["status"] = "completed"
            state["phase_message"] = "Batch parse completed"
            state["completed_at"] = _current_timestamp()
            state["artifacts"] = {
                "summary": cls._summary_path(job_id, root_path).name,
                "results": results_path.name,
                "errors": errors_path.name,
                "excel": excel_path.name if excel_path is not None else None,
            }
            cls._finalize_summary(job_id, state, templates, root_path)
            cls._write_state(job_id, state, root_path)
        except Exception as exc:  # pragma: no cover - failure path
            state["status"] = "failed"
            state["phase_message"] = "Batch parse failed"
            state["completed_at"] = _current_timestamp()
            state["recent_error"] = {
                "error": str(exc),
                "error_type": type(exc).__name__,
                "traceback": traceback.format_exc(),
            }
            state["artifacts"] = {
                "summary": cls._summary_path(job_id, root_path).name,
                "results": cls._results_path(job_id, root_path).name if cls._results_path(job_id, root_path).exists() else None,
                "errors": cls._errors_path(job_id, root_path).name if cls._errors_path(job_id, root_path).exists() else None,
                "excel": cls._excel_path(job_id, root_path).name if cls._excel_path(job_id, root_path).exists() else None,
            }
            cls._finalize_summary(job_id, state, templates, root_path)
            cls._write_state(job_id, state, root_path)
        finally:
            with cls._lock:
                cls._job_threads.pop(job_id, None)
