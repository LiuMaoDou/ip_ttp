import asyncio
import io
import json
import sys
import zipfile

import httpx
from openpyxl import load_workbook

sys.path.insert(0, "../..")

from backend.app.main import app


TEMPLATES = [
    {
        "id": "interfaces",
        "name": "interfaces",
        "template": "<group>\ninterface {{ interface }}\n ip address {{ ip }}/{{ mask }}\n</group>",
    }
]


def rows_to_dicts(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    return [
        {header: value for header, value in zip(headers, row)}
        for row in rows[1:]
    ]


async def wait_for_job_completion(client: httpx.AsyncClient, job_id: str) -> dict:
    for _ in range(100):
        response = await client.get(f"/api/parse/batch/jobs/{job_id}")
        assert response.status_code == 200
        payload = response.json()
        if payload["status"] in {"completed", "failed"}:
            return payload
        await asyncio.sleep(0.05)

    raise AssertionError("Batch parse job did not finish in time")


def test_batch_parse_api_processes_plain_files(tmp_path, monkeypatch):
    monkeypatch.setenv("TTP_BATCH_JOBS_PATH", str(tmp_path / "batch-jobs"))

    async def run_test():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            response = await client.post(
                "/api/parse/batch/jobs",
                data={"templates_json": json.dumps(TEMPLATES)},
                files=[
                    ("files", ("one.txt", b"interface Lo0\n ip address 1.1.1.1/32\n", "text/plain")),
                    ("files", ("two.txt", b"interface Lo1\n ip address 2.2.2.2/32\n", "text/plain")),
                ],
            )

            assert response.status_code == 200
            created = response.json()
            assert created["template_count"] == 1
            assert created["upload_count"] == 2

            completed = await wait_for_job_completion(client, created["id"])
            assert completed["status"] == "completed"
            assert completed["discovered_file_count"] == 2
            assert completed["total_tasks"] == 2
            assert completed["success_count"] == 2
            assert completed["failure_count"] == 0
            assert completed["artifact_urls"]["summary"] is not None
            assert completed["artifact_urls"]["results"] is not None
            assert completed["artifact_urls"]["excel"] is not None

            results_response = await client.get(f"/api/parse/batch/jobs/{created['id']}/results")
            assert results_response.status_code == 200
            results_page = results_response.json()
            assert results_page["total"] == 2
            assert len(results_page["items"]) == 2
            assert all(item["success"] is True for item in results_page["items"])

            summary_response = await client.get(completed["artifact_urls"]["summary"])
            assert summary_response.status_code == 200
            summary = summary_response.json()
            assert summary["counts"]["success_count"] == 2

    asyncio.run(run_test())


def test_batch_parse_api_accepts_camel_case_variable_names(tmp_path, monkeypatch):
    monkeypatch.setenv("TTP_BATCH_JOBS_PATH", str(tmp_path / "batch-jobs"))

    templates = [
        {
            "id": "interfaces",
            "name": "interfaces",
            "template": "<group>\ninterface {{ interface }}\n ip address {{ ip }}/{{ mask }}\n</group>",
            "variableNames": ["interface", "ip", "mask"],
        }
    ]

    async def run_test():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            response = await client.post(
                "/api/parse/batch/jobs",
                data={"templates_json": json.dumps(templates)},
                files=[
                    ("files", ("one.txt", b"interface Lo0\n ip address 1.1.1.1/32\n", "text/plain")),
                ],
            )

            assert response.status_code == 200
            created = response.json()

            completed = await wait_for_job_completion(client, created["id"])
            assert completed["status"] == "completed"
            assert completed["success_count"] == 1
            assert completed["failure_count"] == 0

    asyncio.run(run_test())


def test_batch_parse_excel_artifact_flattens_result_only(tmp_path, monkeypatch):
    monkeypatch.setenv("TTP_BATCH_JOBS_PATH", str(tmp_path / "batch-jobs"))

    templates = [
        {
            "id": "interfaces",
            "name": "Core/Interfaces",
            "template": (
                "<group name=\"device\">\n"
                "hostname {{ hostname }}\n"
                "</group>\n"
                "<group name=\"interfaces*\">\n"
                "interface {{ name }}\n"
                " ip address {{ ip }}/{{ mask }}\n"
                "</group>"
            ),
        },
        {
            "id": "routes",
            "name": "Core?Interfaces",
            "template": (
                "<group name=\"routes*\">\n"
                "ip route {{ prefix }} {{ next_hop }}\n"
                "</group>"
            ),
        },
    ]

    async def run_test():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            response = await client.post(
                "/api/parse/batch/jobs",
                data={"templates_json": json.dumps(templates)},
                files=[
                    (
                        "files",
                        (
                            "edge.txt",
                            (
                                b"hostname edge-1\n"
                                b"interface Lo0\n ip address 1.1.1.1/32\n"
                                b"interface Lo1\n ip address 2.2.2.2/32\n"
                                b"ip route 0.0.0.0/0 10.0.0.1\n"
                            ),
                            "text/plain",
                        ),
                    ),
                ],
            )

            assert response.status_code == 200
            created = response.json()
            completed = await wait_for_job_completion(client, created["id"])
            assert completed["status"] == "completed"
            assert completed["artifact_urls"]["excel"] is not None

            excel_response = await client.get(completed["artifact_urls"]["excel"])
            assert excel_response.status_code == 200
            assert excel_response.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            workbook = load_workbook(filename=io.BytesIO(excel_response.content), data_only=True)
            assert workbook.sheetnames == ["Core_Interfaces", "Core_Interfaces (2)"]

            interfaces_sheet = workbook["Core_Interfaces"]
            interfaces_rows = rows_to_dicts(interfaces_sheet)
            assert interfaces_rows == [
                {
                    "device.hostname": "edge-1",
                    "interfaces.name": "Lo0",
                    "interfaces.ip": "1.1.1.1",
                    "interfaces.mask": "32",
                },
                {
                    "device.hostname": "edge-1",
                    "interfaces.name": "Lo1",
                    "interfaces.ip": "2.2.2.2",
                    "interfaces.mask": "32",
                },
            ]

            routes_sheet = workbook["Core_Interfaces (2)"]
            assert rows_to_dicts(routes_sheet) == [
                {
                    "routes.prefix": "0.0.0.0/0",
                    "routes.next_hop": "10.0.0.1",
                },
            ]

    asyncio.run(run_test())


def test_batch_parse_api_expands_zip_archives_on_server(tmp_path, monkeypatch):
    monkeypatch.setenv("TTP_BATCH_JOBS_PATH", str(tmp_path / "batch-jobs"))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as archive:
        archive.writestr("configs/edge-1.txt", "interface Lo9\n ip address 9.9.9.9/32\n")
        archive.writestr("configs/ignore.bin", b"\x00\x01\x02")

    async def run_test():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
            response = await client.post(
                "/api/parse/batch/jobs",
                data={"templates_json": json.dumps(TEMPLATES)},
                files=[
                    ("files", ("batch.zip", zip_buffer.getvalue(), "application/zip")),
                ],
            )

            assert response.status_code == 200
            created = response.json()
            completed = await wait_for_job_completion(client, created["id"])

            assert completed["status"] == "completed"
            assert completed["upload_count"] == 1
            assert completed["discovered_file_count"] == 1
            assert completed["skipped_file_count"] == 1
            assert completed["total_tasks"] == 1
            assert completed["success_count"] == 1

            results_response = await client.get(f"/api/parse/batch/jobs/{created['id']}/results")
            assert results_response.status_code == 200
            item = results_response.json()["items"][0]
            assert item["file_name"] == "batch.zip::configs/edge-1.txt"
            assert item["success"] is True

    asyncio.run(run_test())
